# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, Border, Color, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from datetime import datetime, time, date, timedelta
import locale

roweek = NamedStyle(name="roweek")
roweek.font = Font(color='FFFFFF', size=12, bold=True)
roweek.fill = PatternFill("solid", fgColor="C5D9F1")
roweek.alignment = Alignment(horizontal='center')

roweek2 = NamedStyle(name="roweek2")
roweek2.font = Font(color='FFFFFF', size=12, bold=True)
roweek2.fill = PatternFill("solid", fgColor="DCE6F1")
roweek2.alignment = Alignment(horizontal='center')

monthPair = NamedStyle(name="monthPair")
monthPair.font = Font(color='FFFFFF', size=12, bold=True)
monthPair.fill = PatternFill("solid", fgColor="BFBFBF")
monthPair.alignment = Alignment(horizontal='center')

monthOdd = NamedStyle(name="monthOdd")
monthOdd.font = Font(color='FFFFFF', size=12, bold=True)
monthOdd.fill = PatternFill("solid", fgColor="D9D9D9")
monthOdd.alignment = Alignment(horizontal='center')

total_week = NamedStyle(name="total_week")
total_week.font = Font(color='FFFFFF', size=12, bold=True)
total_week.fill = PatternFill("solid", fgColor="6E6E6E")
total_week.alignment = Alignment(horizontal='center')


def create_workbook(title):
	wb = Workbook()
	ws = wb.active
	ws.title = title
	return [wb, ws]

def header(ws, enc, h1, h2, filters):
	ws['A1'] = enc
	title = ws['A1']
	title = format_text(ws['A1'], "center", "center")
	ws.merge_cells('A1:K1')

	ws['A3'] = h1
	ws['B3'] = h2
	title = ws['A3']
	title = format_text(ws['A3'], "left", "center")

	init_date  = filters.get('init_date', False)
	last_date  = filters.get('last_date', False)
	canal_est = filters.get('canal_est', False)
	canal_giro = filters.get('canal_giro', False)
	presale_route  = filters.get('presale_route', False)
	delivery_route  = filters.get('delivery_route', False)
	product = filters.get('product', False)
	group_product = filters.get('group_product',False)
	branch = filters.get('branch', False)
	population = filters.get('population',False)
	business_name = filters.get('business_name',False)

	nro = 3
	if init_date and last_date:
		nro += 1
		ws['A'+str(nro)] = 'FECHA'
		ws['B'+str(nro)] = 'Desde '+init_date+' Al '+last_date
		title = ws['A'+str(nro)]
		title = format_text(ws['A'+str(nro)], "left", "center")

	if presale_route:
		if presale_route["presale_route_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'RUTA PREVENTA'
			ws['B'+str(nro)] = str(presale_route["presale_route"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if delivery_route:
		if delivery_route["delivery_route_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'RUTA ENTREGA'
			ws['B'+str(nro)] = str(delivery_route["delivery_route"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if product:
		if product["product_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'PRODUCTO'
			ws['B'+str(nro)] = str(product["product"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if canal_est:
		if canal_est["canal_est_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'DESCRIPCION CANAL'
			ws['B'+str(nro)] = str(canal_est["canal_est"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if canal_giro:
		if canal_giro["canal_giro_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'SEGMENTO'
			ws['B'+str(nro)] = str(canal_giro["canal_giro"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")  
	
	if group_product:
		if group_product["group_product_id"] != 0:
			nro += 1
			ws['A'+str(nro)] = 'GRUPO'
			ws['B'+str(nro)] = str(group_product["group_product"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")
	
	if branch:
		if branch["branch_id"] != 0:
			nro +=1
			ws['A'+str(nro)] = 'MARCA'
			ws['B'+str(nro)] = str(branch["branch"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if population:
		if population["population_id"] != 0:
			nro +=1
			ws['A'+str(nro)] = 'POBLACION'
			ws['B'+str(nro)] = str(population["population"])
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	if business_name:
		if business_name != "":
			nro +=1
			ws['A'+str(nro)] = 'NEGOCIO'
			ws['B'+str(nro)] = str(business_name)
			title = ws['A'+str(nro)]
			title = format_text(ws['A'+str(nro)], "left", "center")

	nro += 1
	return [ws,nro]

def format_text(celda, alingH, alingV):
	title = celda
	title.font = Font(size=12,bold=True)
	title.alignment = Alignment(horizontal=alingH, vertical=alingV)

def resize_cells(ws, size):
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = size
	for col, value in dims.items():
		ws.column_dimensions[col].width = value
	return ws

def week_header(ws, start, row, arrWeeks, jumps):
	for i in range(0,len(arrWeeks)):
		if i == 0:
			start +=1
		else:
			start = end + 1
		end = (start + jumps) - 1

		col_ini = get_column_letter(start)
		col_fin = get_column_letter(end)

		ws.merge_cells(col_ini + str(row) + ':' + col_fin + str(row))

		ws[col_ini + str(row)] = arrWeeks[i]["text"]
		title = ws[col_ini + str(row)]
		if(i % 2 != 0):
			ws[col_ini + str(row)].style = roweek2
		else:
			ws[col_ini + str(row)].style = roweek
		title = format_text(ws[col_ini + str(row)], "center", "center")

	return ws

def month_header(ws, start, row, arrMonths, jumps):

	for i in range(0,len(arrMonths)):
		if i == 0:
			start +=1
		else:
			start = end + 1
		end = (start + (13 * int(arrMonths[i]["weeks"]))) - 1

		col_ini = get_column_letter(start)
		col_fin = get_column_letter(end)

		ws.merge_cells(col_ini + str(row) + ':' + col_fin + str(row))

		ws[col_ini + str(row)] = arrMonths[i]["text"]
		title = ws[col_ini + str(row)]
		if(i % 2 != 0):
			ws[col_ini + str(row)].style = monthOdd
		else:
			ws[col_ini + str(row)].style = monthPair

		title = format_text(ws[col_ini + str(row)], "center", "center")

	return ws

def month_header_array(ws, start, row, arrMonths):

	for i in range(0,len(arrMonths)):
		if i == 0:
			start +=1
		else:
			start = end + 1
		end = (start + arrMonths[i]["cols"]) - 1

		col_ini = get_column_letter(start)
		col_fin = get_column_letter(end)

		ws.merge_cells(col_ini + str(row) + ':' + col_fin + str(row))

		ws[col_ini + str(row)] = arrMonths[i]["text"]
		title = ws[col_ini + str(row)]
		if(i % 2 != 0):
			ws[col_ini + str(row)].style = monthOdd
		else:
			ws[col_ini + str(row)].style = monthPair

		title = format_text(ws[col_ini + str(row)], "center", "center")

	return ws

def week_header_array(ws, start, row, arrWeeks):
	for i in range(0,len(arrWeeks)):
		if i == 0:
			start +=1
		else:
			start = end + 1
		end = (start + arrWeeks[i]["cols"]) - 1

		col_ini = get_column_letter(start)
		col_fin = get_column_letter(end)

		ws.merge_cells(col_ini + str(row) + ':' + col_fin + str(row))

		ws[col_ini + str(row)] = arrWeeks[i]["text"]
		title = ws[col_ini + str(row)]
		if(i % 2 != 0):
			ws[col_ini + str(row)].style = roweek2
		else:
			ws[col_ini + str(row)].style = roweek
		title = format_text(ws[col_ini + str(row)], "center", "center")

	return ws

def load_rows(ws, data):
	for row in data:
		row_list = list(row)
		ws.append(row_list)
	return ws

def freeze_row(ws,num_col,num_fil):
	num_fil += 2
	ws.freeze_panes = ws[num_col + str(num_fil)]

	return ws

def paint_total(ws, longHeader, lenData, arrTotales, row=11):
	roweek.font = Font(color="000000", size=12, bold=True)
	roweek2.font = Font(color="000000", size=12, bold=True)
	roweek.alignment = Alignment(horizontal='right')
	roweek2.alignment = Alignment(horizontal='right')
	for colTotal in range(0,len(arrTotales)):
		column_letter = get_column_letter(arrTotales[colTotal])
		for rowD in range(row,lenData+row+1):
			if(rowD % 2 == 0):
				ws[column_letter + str(rowD)].style = roweek
			else:
				ws[column_letter + str(rowD)].style = roweek2
			ws[column_letter + str(rowD)].number_format = '#,##0.00'
	return ws

def paint_par(ws, longHeader, lenData, num_col,row=11, col_money=[],col_porc=[],col_nro_dec=[],col_nro_int=[]):
	rowPar = NamedStyle(name="rowPar")
	rowPar.fill = PatternFill("solid", fgColor="E0ECF8")

	for column in range(1,longHeader+1):
		column_letter = get_column_letter(column)
		for rowD in range(row,lenData+row+1):
			if(rowD % 2 == 0):
				ws[column_letter + str(rowD)].style = rowPar
			if(column > num_col):
				ws[column_letter + str(rowD)].number_format = '#,##0.00'
			if(column in col_nro_int):
				ws[column_letter + str(rowD)].number_format = '#0'
			if(column in col_nro_dec):
				ws[column_letter + str(rowD)].number_format = '#,##0.00'
			if(column in col_money):
				ws[column_letter + str(rowD)].number_format = '#,##0.00 $'
			if(column in col_porc):
				ws[column_letter + str(rowD)].number_format = '#,##0.00 %'
	return ws

def paint_columns(ws,longHeader, lenData, num_col,row = 11):
	rowPar2 = NamedStyle(name="rowPar2")
	rowPar2.fill = PatternFill("solid", fgColor="E0F8F1")

	ini_col = num_col + 1
	fin_col = ini_col + 13

	while fin_col < longHeader:
		for column in range(ini_col,fin_col):
			column_letter = get_column_letter(column)
			for rowD in range(row,lenData+row+1):
				if(rowD % 2 == 0):
					ws[column_letter + str(rowD)].style = rowPar2

		ini_col = fin_col + 13
		fin_col = ini_col + 13

	rowTotal = NamedStyle(name="rowTotal")
	rowTotal.fill = PatternFill("solid", fgColor="E6E6E6")

	ult_cols = longHeader - 13
	for column in range(ult_cols+1,longHeader+1):
		column_letter = get_column_letter(column)
		for rowD in range(row,lenData+row+1):
			if(rowD % 2 == 0):
				ws[column_letter + str(rowD)].style = rowTotal

	return ws

def load_filters(ws, init_vector):
	FullRange = init_vector +':' + get_column_letter(ws.max_column) \
	+ str(ws.max_row)
	ws.auto_filter.ref = FullRange
	return ws

def total_summary(ws, listTotal, numberRow, lonTableHeader, font_color="FFFFFF", fill_color="afbcd7",formatPercent=[],formatMoney=[],formatNumberInteger=[],formatNumberDecimal=[]):

	ws.append(listTotal)

	totalOpe = NamedStyle(name="totalOpe")
	totalOpe.alignment = Alignment(horizontal='center')
	totalOpe.fill = PatternFill("solid", fgColor=fill_color)
	totalOpe.font = Font(color=font_color, size=12, bold=True)
	for row in ws.iter_rows('A'+str(numberRow)+':'+get_column_letter(lonTableHeader)+str(numberRow)):
		for cell in row:
			cell.style = totalOpe

	for col_letter in formatPercent:
		ws[col_letter+str(numberRow)].number_format = '#,##0.00 %'
	for col_letter in formatMoney:
		ws[col_letter+str(numberRow)].number_format = '#,##0.00 $'
	for col_letter in formatNumberDecimal:
		ws[col_letter+str(numberRow)].number_format = '#,##0.00'
	for col_letter in formatNumberInteger:
		ws[col_letter+str(numberRow)].number_format = '#0'

	return ws

def adds_title_format_new(ws, lonTableHeader, font_color="FFFFFF", fill_color="afbcd7",rows=10):
	headOpe = NamedStyle(name="headOpe")
	headOpe.alignment = Alignment(horizontal='center')
	headOpe.fill = PatternFill("solid", fgColor=fill_color)
	headOpe.font = Font(color=font_color, size=12, bold=True)
	for row in ws.iter_rows('A'+str(rows)+':'+get_column_letter(lonTableHeader)+str(rows)):
		for cell in row:
			cell.style = headOpe
	return ws

def adds_title_format(ws, lonTableHeader, font_color="FFFFFF", fill_color="afbcd7",rows=10):
	headOpe = NamedStyle(name="headOpe")
	headOpe.alignment = Alignment(horizontal='center')
	headOpe.fill = PatternFill("solid", fgColor=fill_color)
	headOpe.font = Font(color=font_color, size=12, bold=True)
	ws.insert_rows(rows)
	rows = rows + 1
	for row in ws.iter_rows('A'+str(rows)+':'+get_column_letter(lonTableHeader)+str(rows)):
		for cell in row:
			cell.style = headOpe
	return ws

def format_date(_date, time):
	date_time = _date + " " + time
	return date_time

def daysBetweenDates(dateStart, dateEnd):
	locale.setlocale(locale.LC_TIME, 'esp')
	arrDates = []
	d = dateStart - timedelta(days=1)
	while d<dateEnd:
		d += timedelta(days=1)
		arrDates.append(
			{'day':d.strftime('%d'),
			'month':d.strftime('%m'),
			'year':d.strftime('%Y'),
			'week':d.isocalendar()[1],
			'monthString': d.strftime('%B').capitalize()
			}
		)
	return arrDates
