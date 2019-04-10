# -*- coding: utf-8 -*-

import os
import sys
import psycopg2
from datetime import datetime, time, date

from flask import Flask, send_file, request, jsonify, json
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from app import app
from app.models import model_client_ranking_week

from app.utils import utils

@app.route('/api/v2/ranking_client_week', methods=['POST'])
def ranking_client_week(charset='utf-8'):
  jsonResponse = json.loads(request.data)
  filters = jsonResponse.get('filters', False)
  # data,week_start,week_end = model_client_ranking_week.ranking_week(filters)
  data, arrDates = model_client_ranking_week.ranking_week(filters)
  wb, ws = utils.create_workbook("Vtas Cltes Rankin x Semana")
  ws,row = utils.header(ws, "Vtas Cltes rankin x Semana Detallada Filtros", 'CEDIS', 'LAGOS DE MORENO',filters)

  # print(jsonResponse.get('table', False))
  # table_exists = data.get('table', False)
  # headers_exists = data['table'].get('table', False)
  # if table_exists and headers_exists:

  arrWeek = []
  for day in arrDates:
    arrWeek.append(day["week"])
  arrWeek1 = set(arrWeek)
  defW = list(arrWeek1)
  # print(defW[len(defW)-1])

  week_start = defW[0]
  week_end = defW[len(defW)-1]

  table_header = jsonResponse['table']['headers']
  start = len(table_header) - 13
  sub_header = table_header[start:]
  for week in range(week_start,week_end):
    table_header = table_header + sub_header
  table_header = table_header + sub_header

  ws = utils.week_header(ws, start, row, defW)
  # print(sub_header)
  # else:
  #   response = { 'response': 'El header de la tabla es requerido'}
  #   return jsonify(response)

  if len(data)>0:
    ws.append(list(table_header))
    # Method for load rows
    ws = utils.load_rows(ws, data)
    ws = utils.freeze_row(ws,'H',row)

    col_porc=[]
    col_money=[]
    col_nro_int = []
    col_nro_dec = []
    col_nro = 7
    for week in range(week_start, week_end+1):
      col_nro += 1
      col_nro_dec.append(col_nro)
      col_nro += 1
      col_porc.append(col_nro)
      for t in range(8):
        col_nro += 1
        col_money.append(col_nro)
      col_nro += 1
      col_nro_int.append(col_nro)
      col_nro += 1
      col_nro_int.append(col_nro)
      col_nro += 1
      col_porc.append(col_nro)

    longData = len(data)
    longHeader = len(table_header)
    ws = utils.paint_par(ws, longHeader, longData, 7,row, col_money, col_porc,col_nro_dec,col_nro_int)
    ws = utils.paint_columns(ws,longHeader, longData, 7,row)
    # Method for rezise cells
    # This method recives workbook active instance and
    ws = utils.resize_cells(ws, 20)
    # Method for load filters
    # This method recives workbook active instance and
    # init vector on init the header_table
    # ws = utils.load_filters(ws, 'A11')
    # Method for adds color to titles, first color si for the font
    # second color is for fill cell. Colors is in format RGB
    ws = utils.adds_title_format(ws, longHeader, "FFFFFF", "4F81BD",row)

    total_start = row + 2
    total_end = (len(data) + total_start) -1
    total_total = total_end + 1

    ini_cant=7
    listTotal = ['Total','','','','','','']
    formatPercent = []
    formatMoney = []
    formatNumberInteger = []
    formatNumberDecimal = []

    for week in range(week_start, week_end+2):
      ini_cant += 1
      letter_col = get_column_letter(ini_cant)
      listTotal.append('= SUM('+letter_col+str(total_start)+':'+letter_col+str(total_end)+')')
      formatNumberDecimal.append(letter_col)

      ini_cant += 1
      letter_col = get_column_letter(ini_cant)
      listTotal.append('= SUM('+letter_col+str(total_start)+':'+letter_col+str(total_end)+')')
      formatPercent.append(letter_col)

      for t in range(8):
        ini_cant += 1
        letter_col = get_column_letter(ini_cant)
        listTotal.append('= SUM('+letter_col+str(total_start)+':'+letter_col+str(total_end)+')')
        formatMoney.append(letter_col)

      ini_cant += 1
      letter_col = get_column_letter(ini_cant)
      listTotal.append('= SUM('+letter_col+str(total_start)+':'+letter_col+str(total_end)+')')
      formatNumberInteger.append(letter_col)
      letter_r = letter_col

      ini_cant += 1
      letter_col = get_column_letter(ini_cant)
      listTotal.append('= SUM('+letter_col+str(total_start)+':'+letter_col+str(total_end)+')')
      formatNumberInteger.append(letter_col)
      letter_s = letter_col

      ini_cant += 1
      letter_col = get_column_letter(ini_cant)
      listTotal.append('= (('+letter_s+str(total_total)+')/'+letter_r+str(total_total)+')')
      formatPercent.append(letter_col)

    ws = utils.total_summary(ws, listTotal, total_total, len(table_header), "FFFFFF", "4F81BD", formatPercent,formatMoney,formatNumberInteger,formatNumberDecimal)

    nombre_archivo = datetime.now().date().strftime('%Y %m %d')+" 18Vtas Cltes rankin x Semana Detallada Filtros.xlsx"
    wb.save(nombre_archivo)
    return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)
  else:
    response = { 'response': 'No hay registros'}
    return jsonify(response)
