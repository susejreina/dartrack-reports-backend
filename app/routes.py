import os
import sys
import psycopg2
from datetime import datetime, time, date

from flask import Flask, send_file, request
from openpyxl.styles import colors, Border, Color, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

from app import app

@app.route('/report/ventas_cliente',methods=['POST'])
def ventas_cliente():
  date_start = request.form['date_start']+' 00:00:00'
  date_end = request.form['date_end']+' 23:59:59'

  headOpe = NamedStyle(name="headOpe")
  headOpe.alignment = Alignment(horizontal='center')
  headOpe.font = Font(color='FFFFFF')

  rowPar = NamedStyle(name="rowPar")
  rowPar.fill = PatternFill("solid", fgColor="E0ECF8")
  
  name_db = os.environ["DATABASE"]
  user_db = os.environ["USER"]
  pass_db = os.environ["PASS"]
  host_db = os.environ["HOST"]

  conn = psycopg2.connect(database=name_db,user=user_db,password=pass_db, host=host_db)
  cur = conn.cursor()
  query = """SELECT O.client_id as id, C.company_code as id_decsa, C.client_number_ant as num_cliente, C.client_high as cliente_mayorista, C.business_name as negocio, CE.name as cedis, R.name as route, CH.name as channel, CT.name as segmento, CD.gec_name as gec, CD.condition_name as condition
            FROM orders O 
            LEFT JOIN clients C
            ON O.client_id = C.id
            LEFT JOIN routes R
            ON O.route_id = R.id
            LEFT JOIN centers CE
            ON O.center_id = CE.id
            LEFT JOIN channels CH
            ON C.channel_id = CH.id
            LEFT JOIN client_types CT
            ON C.client_type_id = CT.id
            LEFT JOIN client_details CD
            ON CD.client_id = C.id
            WHERE O.id IN
            (
            SELECT MAX(id) as id
            FROM orders
            WHERE active=true AND ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
            GROUP BY client_id
            ORDER BY client_id
            )
            ORDER BY O.client_id;"""
  cur.execute(query)
  data = cur.fetchall()
  query = """SELECT	o.client_id, EXTRACT(MONTH FROM o.ordered_at) as month, EXTRACT(YEAR FROM o.ordered_at) as year, SUM(od.quantity_delivered)::INTEGER "boxes", SUM(od.hlts)::DOUBLE PRECISION "htls"
            FROM order_details od
            INNER JOIN orders  o ON o.id = od.order_id
            INNER JOIN products  p ON p.id = od.product_id
            WHERE is_devolution = false AND o.active = true AND o.ordered_at 
            BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
            GROUP BY O.client_id, EXTRACT(MONTH FROM O.ordered_at), EXTRACT(YEAR FROM O.ordered_at)
            ORDER BY O.client_id, EXTRACT(YEAR FROM O.ordered_at), EXTRACT(MONTH FROM O.ordered_at);"""
  cur.execute(query)
  data_sum = cur.fetchall()

  wb = Workbook()
  ws = wb.active
  ws.title = "Clte Acum Ano"

  yearStart = int(date_start[0:4])
  monthStart = int(date_start[5:7])
  yearEnd = int(date_end[0:4])
  monthEnd = int(date_end[5:7])

  months = [
    [1,"Ene"],
    [2,"Feb"],
    [3,"Mar"],
    [4,"Abr"],
    [5,"May"],
    [6,"Jun"],
    [7,"Jul"],
    [8,"Ago"],
    [9,"Sep"],
    [10,"Oct"],
    [11,"Nov"],
    [12,"Dic"],
    [100,"Total"],
  ]
  qtyMonths = len(months)

  ws['A1'] = "80-DECSA Clte Acum año mes solo Cjs Htls"

  title = ws['A1']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="center", vertical="center")
  ws.merge_cells('A1:K1')

  ws['A3'] = 'CEDIS'
  ws['B3'] = 'LAGOS DE MORENO'
  title = ws['A3']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  ws['A4'] = 'FECHA'
  ws['B4'] = 'Desde '+request.form['date_start']+' Al '+request.form['date_end']
  title = ws['A4']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")
  if len(data)>0:
    cell_hasta = len(data) + 6
    # add column headings. NB. these must be strings
    table_header = ["Id", "Id Decsa", "Num. Cliente", "Cve Cte Mayorista", "Negocio", "Cedis", "Ruta de Venta", "Canal Estrat.", "Segmento", "GEC", "Condición"]
    begin = yearStart
    end = int(yearEnd)+1
    for y in range(begin,end):
      if y == yearStart and y == yearEnd:
        mStart = monthStart - 1
        mEnd = monthEnd
      elif y == yearStart:
        mStart = monthStart - 1
        mEnd = qtyMonths
      elif y == yearEnd:
        mStart = 0
        mEnd = monthEnd
      else:
        mStart==0
        mEnd = qtyMonths
      print('hello----')
      print(mStart)
      for mm in range(mStart,mEnd):
        table_header.append(months[mm][1]+" "+str(y)+"\nHlts")
    if mEnd<qtyMonths:
      table_header.append(months[qtyMonths-1][1]+" "+str(y)+"\nHlts")
    htls_end = len(table_header)
    box_start = htls_end + 1
    for y in range(begin,end):
      if y == yearStart and y == yearEnd:
        mStart = monthStart - 1
        mEnd = monthEnd
      elif y == yearStart:
        mStart = monthStart - 1
        mEnd = qtyMonths
      elif y == yearEnd:
        mStart = 0
        mEnd = monthEnd
      else:
        mStart==0
        mEnd = qtyMonths
      for mm in range(mStart,mEnd):
        table_header.append(months[mm][1]+" "+str(y)+"\nCajas")
    if mEnd<qtyMonths:
      table_header.append(months[qtyMonths-1][1]+" "+str(y)+"\nCajas")
    ws.append(table_header)

    box_end = len(table_header)
    row_ds = 0
    for row in data:
      row_list = list(row)
      for y in range(begin,end):
        total = 0
        if y == yearStart and y == yearEnd:
          mStart = monthStart - 1
          mEnd = monthEnd
        elif y == yearStart:
          mStart = monthStart - 1
          mEnd = qtyMonths
        elif y == yearEnd:
          mStart = 0
          mEnd = monthEnd
        else:
          mStart==0
          mEnd = qtyMonths
        for mm in range(mStart,mEnd):
          valor = 0
          for ind in range(row_ds,len(data_sum)):
            if (data_sum[ind][0]==row[0] and data_sum[ind][1]==months[mm][0] and data_sum[ind][2]==y):
              row_ds = ind + 1
              valor = data_sum[ind][4]
              total += valor
              break
          if int(months[mm][0])==100:
            valor = total
          row_list.append(valor)
        if mEnd<qtyMonths:
          row_list.append(total)
      row_ds = 0
      for y in range(begin,end):
        total = 0
        if y == yearStart and y == yearEnd:
          mStart = monthStart - 1
          mEnd = monthEnd
        elif y == yearStart:
          mStart = monthStart - 1
          mEnd = qtyMonths
        elif y == yearEnd:
          mStart = 0
          mEnd = monthEnd
        else:
          mStart==0
          mEnd = qtyMonths
        for mm in range(mStart,mEnd):
          valor = 0
          for ind in range(row_ds,len(data_sum)):
            if (data_sum[ind][0]==row[0] and data_sum[ind][1]==months[mm][0] and data_sum[ind][2]==y):
              row_ds = ind + 1
              valor = data_sum[ind][3]
              total += valor
              break
          if int(months[mm][0])==100:
            valor = total
          row_list.append(valor)
        if mEnd<qtyMonths:
          row_list.append(total)
      ws.append(row_list)

    dims = {}
    for row in ws.rows:
      for cell in row:
        if cell.value:
          dims[cell.column] = 15
    for col, value in dims.items():
      ws.column_dimensions[col].width = value

    ws.insert_rows(5)
    title = ws['A7']
    ws.freeze_panes = title

    for column in range(1,len(table_header)+1):
      column_letter = get_column_letter(column)
      for rowD in range(5,len(data)+6):
        if(rowD % 2 == 0):
          ws[column_letter + str(rowD)].style = rowPar

    for row in ws.iter_rows("L6:"+get_column_letter(htls_end)+str(cell_hasta)):
      for cell in row:
        cell.number_format = '#,#0.0'

    tab = Table(displayName="Table1",ref="A6:"+get_column_letter(len(table_header))+str(cell_hasta))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    for row in ws.iter_rows('A6:O6'):
      for cell in row:
        cell.style = headOpe

    title = ws['L7']
    ws.freeze_panes = title

    ws.merge_cells('L5:'+get_column_letter(htls_end)+'5')
    ws['L5'] = 'HLTS'
    hlts = ws['L5']
    hlts.font = Font(color=colors.WHITE,name='Tahoma',size=14)
    hlts.alignment = Alignment(horizontal='center')
    hlts.fill = PatternFill("solid", fgColor="305496")

    ws.merge_cells(get_column_letter(box_start)+'5:'+get_column_letter(box_end)+'5')
    ws[get_column_letter(box_start)+'5'] = 'CAJAS'
    hlts = ws[get_column_letter(box_start)+'5']
    hlts.font = Font(color=colors.BLACK,name='Tahoma',size=14)
    hlts.alignment = Alignment(horizontal='center')
    hlts.fill = PatternFill("solid", fgColor="DDDDDD")

    nombre_archivo ="80-Cajas Hlts" + datetime.now().date().strftime('%d-%m-%Y') + ".xlsx"
    wb.save(nombre_archivo)

    return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)
  else:
    return "No hay registros"

@app.route('/report/ventas_marca',methods=['POST'])
def ventas_marca():
  date_start = request.form['date_start']+' 00:00:00'
  date_end = request.form['date_end']+' 23:59:59'

  headOpe = NamedStyle(name="headOpe")
  headOpe.alignment = Alignment(horizontal='center')
  headOpe.font = Font(color='FFFFFF')

  rowPar = NamedStyle(name="rowPar")
  rowPar.fill = PatternFill("solid", fgColor="E0ECF8")

  name_db = os.environ["DATABASE"]
  user_db = os.environ["USER"]
  pass_db = os.environ["PASS"]
  host_db = os.environ["HOST"]

  conn = psycopg2.connect(database=name_db,user=user_db,password=pass_db, host=host_db)
  cur = conn.cursor()
  query = """SELECT O.client_id as id_client, C.company_code as id_erp, C.client_number_ant as cliente_ceveceria, 
          C.client_high as cliente_mayorista, C.business_name as negocio, P.code, P.name
          FROM order_details OD 
          LEFT JOIN orders O ON OD.order_id = O.id
          LEFT JOIN clients C ON O.client_id = C.id
          LEFT JOIN products P ON OD.product_id=P.id
          WHERE P.active= true AND O.active=true AND O.ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
          GROUP BY O.client_id, C.company_code, C.client_number_ant, C.client_high, C.business_name, P.code, P.name
          ORDER BY O.client_id, P.code;"""
  cur.execute(query)
  data = cur.fetchall()
  query1 = """SELECT O.client_id, P.code, CAST (EXTRACT(YEAR FROM O.ordered_at) AS INTEGER), CAST (EXTRACT(MONTH FROM O.ordered_at) AS INTEGER), SUM(OD.hlts) as hlts
          FROM order_details OD 
          LEFT JOIN orders O ON OD.order_id = O.id
          LEFT JOIN products P ON OD.product_id=P.id
          WHERE P.active= true AND O.active=true AND O.ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
          GROUP BY O.client_id, P.code, EXTRACT(YEAR FROM O.ordered_at), EXTRACT(MONTH FROM O.ordered_at)
          ORDER BY O.client_id, P.code, EXTRACT(YEAR FROM O.ordered_at), EXTRACT(MONTH FROM O.ordered_at);"""
  cur.execute(query1)
  data_sum = cur.fetchall()

  wb = Workbook()
  ws = wb.active
  ws.title = "Clte Acum Ano"

  yearStart = int(date_start[0:4])
  monthStart = int(date_start[5:7])
  yearEnd = int(date_end[0:4])
  monthEnd = int(date_end[5:7])

  months = [
    [1,"Ene"],
    [2,"Feb"],
    [3,"Mar"],
    [4,"Abr"],
    [5,"May"],
    [6,"Jun"],
    [7,"Jul"],
    [8,"Ago"],
    [9,"Sep"],
    [10,"Oct"],
    [11,"Nov"],
    [12,"Dic"],
    [100,"Total"],
  ]
  qtyMonths = len(months)

  ws['A1'] = '81-DECSA Vtas Clte Acum Año Mes Producto Cajas Htls'
  title = ws['A1']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="center", vertical="center")
  ws.merge_cells('A1:G1')

  ws['A3'] = 'CEDIS'
  ws['B3'] = 'LAGOS DE MORENO'
  title = ws['A3']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  ws['A4'] = 'FECHA'
  ws['B4'] = 'Desde '+request.form['date_start']+' Al '+request.form['date_end']
  title = ws['A4']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")
  if len(data)>0:
    cell_hasta = len(data) + 6
    # add column headings. NB. these must be strings
    table_header = ["Id Cliente", "Id ERP", "Num. Cliente", "Cve Cte Mayorista", "Negocio", "Clave", "Producto"]
    begin = yearStart
    end = int(yearEnd)+1
    for y in range(begin,end):
      if y == yearStart and y == yearEnd:
        mStart = monthStart - 1
        mEnd = monthEnd
      elif y == yearStart:
        mStart = monthStart - 1
        mEnd = qtyMonths
      elif y == yearEnd:
        mStart = 0
        mEnd = monthEnd
      else:
        mStart==0
        mEnd = qtyMonths
      for mm in range(mStart,mEnd):
        table_header.append(months[mm][1]+" "+str(y)+"\nHlts")
    if mEnd<qtyMonths:
      table_header.append(months[qtyMonths-1][1]+" "+str(y)+"\nHlts")
    ws.append(table_header)
    hlts_end = len(table_header)

    row_ds = 0
    for row in data:
      row_list = list(row)
      cliente = int(row[0])
      producto = int(row[5])
      details = []
      for ind in range(row_ds,len(data_sum)):
        if cliente==int(data_sum[row_ds][0]) and producto==int(data_sum[row_ds][1]):
          details.append([data_sum[row_ds][2],data_sum[row_ds][3],data_sum[row_ds][4]])
          row_ds += 1
        else:
          break
      for y in range(begin,end):
        if y == yearStart and y == yearEnd:
          mStart = monthStart - 1
          mEnd = monthEnd
        elif y == yearStart:
          mStart = monthStart - 1
          mEnd = qtyMonths
        elif y == yearEnd:
          mStart = 0
          mEnd = monthEnd
        else:
          mStart==0
          mEnd = qtyMonths
        total = 0
        for mm in range(mStart,mEnd):
          valor = 0
          if(months[mm][0]==100):
            valor = total
          else:
            if len(details)>0:
              for d in details:
                if y==int(d[0]) and months[mm][0]==int(d[1]):
                  valor = d[2]
                  total += valor
          if int(months[mm][0])==100:
            valor = total
          row_list.append(valor)
        if mEnd<qtyMonths:
          row_list.append(total)
      ws.append(row_list)

    dims = {}
    for row in ws.rows:
      for cell in row:
        if cell.value:
          dims[cell.column] = 15
    for col, value in dims.items():
      ws.column_dimensions[col].width = value

    ws.insert_rows(5)
    title = ws['A7']
    ws.freeze_panes = title

    for column in range(1,len(table_header)+1):
      column_letter = get_column_letter(column)
      for rowD in range(5,len(data)+6):
        if(rowD % 2 == 0):
          ws[column_letter + str(rowD)].style = rowPar
        if(column_letter == 'H' or column_letter == 'I'):
          ws[column_letter + str(rowD)].number_format = '#,#0.0'

    for row in ws.iter_rows("L6:"+get_column_letter(hlts_end)+str(cell_hasta)):
      for cell in row:
        cell.number_format = '#,#0.0'

    tab = Table(displayName="Table1",ref="A6:"+get_column_letter(len(table_header))+str(cell_hasta))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    for row in ws.iter_rows('A6:'+get_column_letter(len(table_header))+'6'):
      for cell in row:
        cell.style = headOpe

    title = ws['H7']
    ws.freeze_panes = title
    
    ws.merge_cells('H5:'+get_column_letter(hlts_end)+'5')
    ws['H5'] = 'HLTS'
    hlts = ws['H5']
    hlts.font = Font(color=colors.WHITE,name='Tahoma',size=14)
    hlts.alignment = Alignment(horizontal='center')
    hlts.fill = PatternFill("solid", fgColor="305496")
    

    nombre_archivo = "81-DECSA Vtas Clte Acum Año Mes Producto Cajas Htls-" + datetime.now().date().strftime('%d-%m-%Y') + ".xlsx"
    wb.save(nombre_archivo)

    return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)
  else:
    return "No hay registros"

@app.route('/report/ventas_cobertura', methods=["POST"])
def ventas_cobertura():
  name_db = os.environ["DATABASE"]
  user_db = os.environ["USER"]
  pass_db = os.environ["PASS"]
  host_db = os.environ["HOST"]

  conn = psycopg2.connect(database=name_db,user=user_db,password=pass_db, host=host_db)
  cur = conn.cursor()
  #Creamos el libro de trabajo
  wb = Workbook()
  #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
  ws = wb.active
  ws.title = "Cobertura Equipos"

  #month = datetime.now().date().strftime('%m')
  #year = datetime.now().date().strftime('%Y')

  date_start = request.form['date_start']+' 00:00:00'
  date_end = request.form['date_end']+' 23:59:59'

  yearStart = int(date_start[0:4])
  monthStart = int(date_start[5:7])
  yearEnd = int(date_end[0:4])
  monthEnd = int(date_end[5:7])

  headOpe = NamedStyle(name="headOpe")
  headOpe.alignment = Alignment(horizontal='center')
  headOpe.font = Font(color='FFFFFF')

  rowPar = NamedStyle(name="rowPar")
  rowPar.fill = PatternFill("solid", fgColor="E0ECF8")

  months = [
    [1,"Ene"],
    [2,"Feb"],
    [3,"Mar"],
    [4,"Abr"],
    [5,"May"],
    [6,"Jun"],
    [7,"Jul"],
    [8,"Ago"],
    [9,"Sep"],
    [10,"Oct"],
    [11,"Nov"],
    [12,"Dic"],
    [100,"Total"]
  ]
  qtyMonths = len(months)

  ws.merge_cells('A1:N1')
  ws['A1'] = '82-DECSA Vtas Clte Cobertura Equipos Promos Fecha Alta'
  title = ws['A1']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="center", vertical="center")

  ws['A3'] = 'Cedis'
  ws['B3'] = 'Lagos de Moreno'
  title = ws['A3']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  ini = date_start.split(" ")
  fin = date_end.split(" ")
  ws['A4'] = 'Fecha'
  ws['B4'] = 'Desde '+ini[0]+' al '+fin[0]
  title = ws['A4']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  query = """SELECT O.client_id as id_client, C.company_code as id_erp, C.client_number_ant as cliente_ceveceria, C.client_high as mayorista,
            	C.business_name as negocio, CI.name as poblacion, CD.channel_type_name as canal, CT.name as segmento, R.name as route,
            	U.name || ' ' || U.lastname as name, CD.margin_level, C.discount as bonificacion, C.client_status, 
              CASE WHEN C.has_cooler = TRUE THEN 1 ELSE 0 END as enfriador,
              CASE WHEN C.rice = TRUE THEN 1 ELSE 0 END  as rice_entregado, 
              C.rice_amount as monto_rice, 
              CASE WHEN C.has_facade = TRUE THEN 1 ELSE 0 END  as fachada, 
              CASE WHEN C.is_completed = TRUE THEN 1 ELSE 0 END  as doc, 
              C.comments as observ,
              to_char(C.created_at,'YYYY/MM/DD') as creado, to_char(C.created_at, 'TMMonth')
            FROM orders O
            LEFT JOIN clients C ON O.client_id = C.id
            LEFT JOIN addresses A ON C.address_id = A.id
            LEFT JOIN cities CI ON A.city_id = CI.id
            LEFT JOIN routes R ON O.route_id = R.id
            LEFT JOIN channels CH ON C.channel_id = CH.id
            LEFT JOIN client_types CT ON C.client_type_id = CT.id
            LEFT JOIN client_details CD ON CD.client_id = C.id
            LEFT JOIN users U ON O.user_id = U.id
            WHERE R.route_type_id=2 AND O.id IN
            (
            SELECT MAX(id) as id
            FROM orders
            WHERE active=true AND ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
            GROUP BY client_id, route_id
            ORDER BY client_id
            )
            ORDER BY O.client_id;"""

  query1 = """SELECT O.client_id, EXTRACT(MONTH FROM O.ordered_at) as month, EXTRACT(YEAR FROM O.ordered_at) as year, SUM(O.hlts)
            FROM orders O
            LEFT JOIN routes R ON O.route_id = R.id
            WHERE O.active=true
            	AND O.ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
            	AND R.route_type_id=2
            GROUP BY O.client_id, EXTRACT(MONTH FROM O.ordered_at), EXTRACT(YEAR FROM O.ordered_at), route_id
            ORDER BY O.client_id,EXTRACT(YEAR FROM O.ordered_at),EXTRACT(MONTH FROM O.ordered_at);"""


  cur.execute(query)
  data = cur.fetchall()

  cur.execute(query1)
  data_sum = cur.fetchall()


  row_ds = 0
  if len(data)>0:
      cell_hasta = len(data) + 7
      table_header = ["Id Cliente", "Id Decsa", "Num. Cliente", "Cve Cte Mayorista", "Negocio", "Poblacion", "Canal", "Segmento", "Ruta Conq", "Nombre Conquistador", "Gpo Nivel Margen", "% Bonificacion", "CL / NUEVO", "Enfriador", "RICE Entregado", "Monto del RICE", "Fachadas", "Docum. Completa", "Observaciones", "Fecha de alta", "MES DE ALTA"];

      begin = yearStart
      end = int(yearEnd)+1

      for y in range(begin,end):
        if y == yearStart and y == yearEnd:
          mStart = monthStart - 1
          mEnd = monthEnd
        elif y == yearStart:
          mStart = monthStart - 1
          mEnd = qtyMonths
        elif y == yearEnd:
          mStart = 0
          mEnd = monthEnd
        else:
          mStart==0
          mEnd = qtyMonths

        for m in range(mStart,mEnd):
            table_header.append(months[m][1]+" "+str(y))

      if mEnd<qtyMonths:
          table_header.append(months[qtyMonths-1][1]+" "+str(y))

      ws.append(table_header)

      for row in data:
        row_list = list(row)
        for y in range(begin,end):
          total = 0
          if y == yearStart and y == yearEnd:
            mStart = monthStart - 1
            mEnd = monthEnd
          elif y == yearStart:
            mStart = monthStart - 1
            mEnd = qtyMonths
          elif y == yearEnd:
            mStart = 0
            mEnd = monthEnd
          else:
            mStart==0
            mEnd = qtyMonths

          for m in range(mStart,mEnd):
            valor = 0
            for ind in range(row_ds,len(data_sum)):
                if (data_sum[ind][0]==row[0] and data_sum[ind][1]==months[m][0] and data_sum[ind][2]==y):
                    row_ds = ind + 1
                    valor = data_sum[ind][3]
                    total = total + valor
                    break
            if int(months[m][0])==100:
                valor = total

            row_list.append((valor))
          if mEnd<qtyMonths:
            row_list.append((total))

        ws.append(row_list)

      for column in range(1,len(table_header)+1):
        column_letter = get_column_letter(column)
        for rowD in range(5,len(data)+6):
         if(rowD % 2 == 0):
            ws[column_letter + str(rowD)].style = rowPar
         if(column==12 or column==16 or column > 21):
            ws[column_letter + str(rowD)].number_format = '#,#0.0'

      #cambiar ancho de columns
      dims = {}
      for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = 20
      for col, value in dims.items():
        ws.column_dimensions[col].width = value

      ws.insert_rows(5)
      ws.merge_cells("V5:"+get_column_letter(len(table_header))+'5')
      ws['V5'] = 'HTL'
      title = ws['V5']
      title.font = Font(size=14,bold=True)
      title.alignment = Alignment(horizontal="center", vertical="center")
      title.fill = PatternFill("solid", fgColor="2E9AFE")

      #title = ws['D7']
      #ws.freeze_panes = title

      tab = Table(displayName="Table1",ref="A6:"+get_column_letter(len(table_header))+str(cell_hasta))

      # Add a default style with striped rows and banded columns
      style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
      tab.tableStyleInfo = style
      ws.add_table(tab)
      for row in ws.iter_rows('A6:'+get_column_letter(len(table_header))+'6'):
        for cell in row:
         cell.style = headOpe

      nombre_archivo = "82-DECSA Vtas Clte Cobertura Equipos Promos Fecha Alta-" + datetime.now().date().strftime('%d-%m-%Y') + ".xlsx"
      wb.save(nombre_archivo)

      return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)
  else:
    return "No hay registros"

@app.route('/report/ventas_operaciones', methods=["POST"])
def ventas_operaciones():
  name_db = os.environ["DATABASE"]
  user_db = os.environ["USER"]
  pass_db = os.environ["PASS"]
  host_db = os.environ["HOST"]

  conn = psycopg2.connect(database=name_db,user=user_db,password=pass_db, host=host_db)
  cur = conn.cursor()
  #Creamos el libro de trabajo
  wb = Workbook()
  #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
  ws = wb.active
  ws.title = "Vtas Acum Ruta"

  date_start = request.form['date_start']+' 00:00:00'
  date_end = request.form['date_end']+' 23:59:59'

  headOpe = NamedStyle(name="headOpe")
  headOpe.alignment = Alignment(horizontal='center')
  headOpe.font = Font(color='FFFFFF')

  rowPar = NamedStyle(name="rowPar")
  rowPar.fill = PatternFill("solid", fgColor="E0ECF8")

  ws['A1'] = '83-DECSA Vtas Acum Ruta Año Sem Poblacion Cjs Hlts'
  title = ws['A1']
  title.font = Font(size=14,bold=True)
  title.alignment = Alignment(horizontal="center", vertical="center")
  ws.merge_cells('A1:G1')

  ws['A3'] = 'Cedis'
  ws['B3'] = 'Lagos de Moreno'
  title = ws['A3']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  ws['A4'] = 'Fecha'
  ws['B4'] = 'Desde '+request.form['date_start']+' Al '+request.form['date_end']
  title = ws['A4']
  title.font = Font(size=12,bold=True)
  title.alignment = Alignment(horizontal="left", vertical="center")

  query = """SELECT CT.name as cedis, EXTRACT(YEAR FROM O.ordered_at) as year,
                EXTRACT(DOW FROM O.ordered_at) as week,
            	CI.name as poblacion, R.name as ruta, RT.name as modalidad,
            	SUM(quantity) as cajas_pedido,  SUM(quantity_delivered) as cajas_entregadas,
            	SUM(P.hlts::numeric * quantity) as htl_pedido,
            	SUM(P.hlts::numeric * quantity_delivered) as htl_entregado,
            	(SUM(quantity_delivered::numeric) / SUM(quantity::numeric)) as eficiencia
            FROM orders O
            LEFT JOIN order_details OD ON O.id = OD.order_id
            LEFT JOIN clients C ON O.client_id = C.id
            LEFT JOIN centers CT ON C.center_id = CT.id
            LEFT JOIN addresses A ON C.address_id = A.id
            LEFT JOIN cities CI ON A.city_id = CI.id
            LEFT JOIN routes R ON O.route_id = R.id
            LEFT JOIN route_types RT ON R.route_type_id = RT.id
            LEFT JOIN client_details CD ON CD.client_id = C.id
            LEFT JOIN products P ON OD.product_id = P.id
            WHERE R.route_type_id=2 AND CT.id = 1 AND P.active= true AND O.id IN
            (
            SELECT MAX(id) as id
            FROM orders
            WHERE active=true AND ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
            GROUP BY client_id
            )
            GROUP BY CT.name, RT.name, EXTRACT(YEAR FROM O.ordered_at), EXTRACT(DOW FROM O.ordered_at), CI.name, R.name
            ORDER BY EXTRACT(YEAR FROM O.ordered_at), EXTRACT(DOW FROM O.ordered_at);"""

  cur.execute(query)
  data = cur.fetchall()

  if len(data)>0:
      cell_hasta = len(data) + 7
      # add column headings. NB. these must be strings
      header = ["Zona", "SEM", "ANO", "Localidad", "Ruta", "Modalidad de venta", "Cajas - pedidas", "Cajas - entregadas", "HTL - Pedidos", "HTL - Entregados", "% de eficiencia"]
      ws.append(header)

      for row in data:
        ws.append(row)

      for column in range(1,len(header)+1):
        column_letter = get_column_letter(column)
        for rowD in range(5,len(data)+6):
         if(rowD % 2 == 0):
            ws[column_letter + str(rowD)].style = rowPar
         if(column_letter == 'I' or column_letter == 'J'):
            ws[column_letter + str(rowD)].number_format = '#,#0.0'
         if(column_letter == 'K'):
            ws[column_letter + str(rowD)].number_format = '0.00%'


      #cambiar ancho de columns
      dims = {}
      for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = 20
      for col, value in dims.items():
        ws.column_dimensions[col].width = value

      ws.insert_rows(5)
      title = ws['B7']
      ws.freeze_panes = title

      tab = Table(displayName="Table1",ref="A6:K"+str(cell_hasta))

      # Add a default style with striped rows and banded columns
      style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
      tab.tableStyleInfo = style
      ws.add_table(tab)
      for row in ws.iter_rows('A6:K6'):
        for cell in row:
         cell.style = headOpe

      nombre_archivo = "83-DECSA Vtas Acum Ruta Año Sem Poblacion Cjs Hlts-" + datetime.now().date().strftime('%d-%m-%Y') + ".xlsx"
      wb.save(nombre_archivo)

      return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)
  else:
      return "No hay registros"

@app.route('/report/trade_mkt', methods=["POST"])
def trade_mkt():
  name_db = os.environ["DATABASE"]
  user_db = os.environ["USER"]
  pass_db = os.environ["PASS"]
  host_db = os.environ["HOST"]

  conn = psycopg2.connect(database=name_db,user=user_db,password=pass_db, host=host_db)
  #conn = psycopg2.connect(database='sicamex_test',user='sicamex',password='sic@mex123', host='cuamoc-dev.sicamex.mx')
  cur = conn.cursor()
  #Creamos el libro de trabajo
  wb = Workbook()
  #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
  ws = wb.active
  ws.title = "MKT Acum Clte"

  date_start = request.form['date_start']+' 00:00:00'
  date_end = request.form['date_end']+' 23:59:59'

  headOpe = NamedStyle(name="headOpe")
  headOpe.alignment = Alignment(horizontal='center')
  headOpe.font = Font(color='FFFFFF')

  rowPar = NamedStyle(name="rowPar")
  rowPar.fill = PatternFill("solid", fgColor="E0ECF8")

  ws['A1'] = '84-DECSA MKT Acum Clte Canal Promo Iniciativas'
  title = ws['A1']
  title.font = Font(size=14,bold=True)
  title.alignment = Alignment(horizontal="center", vertical="center")
  ws.merge_cells('A1:J1')

  query = """SELECT O.client_id as id, C.company_code as id_decsa, C.client_number_ant as num_cliente, C.client_high as cliente_mayorista,
        C.business_name as negocio,  R.name as route, CH.name as channel, CT.name as segmento,  CD.condition_name as condition,
        CONCAT(REPLACE(REPLACE(C.monday::text,'true','1'),'false',''),REPLACE(REPLACE(C.tuesday::text,'true','2'),'false',''),REPLACE(REPLACE(C.wednesday::text,'true','3'),'false',''),REPLACE(REPLACE(C.thursday::text,'true','4'),'false',''),REPLACE(REPLACE(C.friday::text,'true','5'),'false',''),REPLACE(REPLACE(C.saturday::text,'true','6'),'false',''),REPLACE(REPLACE(C.sunday::text,'true','7'),'false','')) as diasvisita,
        C.is_tkt, C.is_pay_for_performance, C.is_hnk, C.is_portapendo,
        C.is_pay_for_performance, C.is_hnk, C.is_upsizing, C.is_pendon
        FROM orders O
        LEFT JOIN clients C ON O.client_id = C.id
        LEFT JOIN routes R ON O.route_id = R.id
        LEFT JOIN channels CH ON C.channel_id = CH.id
        LEFT JOIN client_types CT ON C.client_type_id = CT.id
        LEFT JOIN client_details CD ON CD.client_id = C.id
        WHERE O.id IN
        (
        SELECT MAX(id) as id
        FROM orders
        WHERE active=true AND ordered_at BETWEEN ('"""+date_start+"""') AND ('"""+date_end+"""')
        GROUP BY client_id
        ORDER BY client_id
        )
        ORDER BY O.client_id;"""
  cur.execute(query)
  data = cur.fetchall()

  query = """SELECT C.id,C.channel_id,CO.group
          FROM clients C
          LEFT JOIN client_combos CC
          ON C.id=CC.client_id
          LEFT JOIN combos CO
          ON CC.combo_id=CO.id
          WHERE C.channel_id=1
          ORDER BY C.id;"""
  cur.execute(query)
  dataOnTrade = cur.fetchall()

  query = """SELECT C.id,C.channel_id,CO.group
          FROM clients C
          LEFT JOIN client_combos CC
          ON C.id=CC.client_id
          LEFT JOIN combos CO
          ON CC.combo_id=CO.id
          WHERE C.channel_id=2 AND CO.group is not null
          ORDER BY C.id;"""
  cur.execute(query)
  dataOffTrade = cur.fetchall()  

  if len(data)>0:
    cell_hasta = len(data) + 3
    # add column headings. NB. these must be strings
    header = ["Id", "Id Decsa", "Num. Cliente", "Cve Cte Mayorista", "Negocio", "Ruta Prev Trad", "Canal Estrategico", "Segmento", "Condicion", "Dias de Visita", "Latones","Caguamones","Caguamón CB","Caguamon XX","HNK","Latas 12 oz","Medias", "SOCIO TECATE","PxD","RETO HNK","PORTAPENDON","PB (Medias)","PB (Litro 1/4)","GRAND HNK + TKTL","HNK MEDIA + TKTL","PxD","RETO HNK","UPSIZING","PENDON"]
    ws.append(header)

    for row in data:
      row_row = list(row)
      row1 = []
      for i in range(0,9):
        row1.append(row_row[i])
      row1.append("")
      if len(row_row[9])>0:
        indice = 1
        diaSemana =  row_row[9][0]
        while indice < len(row_row[9]):
          diaSemana = diaSemana+","+row_row[9][indice]
          indice += 1
        row1[9] = diaSemana
      for i in range(10,17):
        row1.append(0)
      for dOT in dataOnTrade:
        if row_row[0]==dOT[0]:
          if dOT[2]=='Latones':
            row1[10]=1
          elif dOT[2]=='Caguamones':
            row1[11]=1
          elif dOT[2]=='Caguamon CB':
            row1[12]=1
          elif dOT[2]=='Caguamon XX':
            row1[13]=1
          elif dOT[2]=='HNK':
            row1[14]=1
          elif dOT[2]=='Latas 12 oz':
            row1[15]=1
          elif dOT[2]=='Medias':
            row1[16]=1
      for i in range(10,14):
        if row_row[i]:
          row1.append(1)
        else:
          row1.append(0)
      for i in range(21,25):
        row1.append(0)
      for dOT in dataOffTrade:
        if row_row[0]==dOT[0]:
          if dOT[2]=='PB (Medias)':
            row1[21]=1
          elif dOT[2]=='"PB (Litro 1/4)':
            row1[22]=1
          elif dOT[2]=='GRAND HNK + TKTL':
            row1[23]=1
          elif dOT[2]=='HNK MEDIA + TKTL':
            row1[24]=1
      for i in range(14,18):
        if row_row[i]:
          row1.append(1)
        else:
          row1.append(0)
      ws.append(row1)
    
    for column in range(1,len(header)+1):
      column_letter = get_column_letter(column)
      for rowD in range(4,len(data)+3):
        if(rowD % 2 == 0):
          ws[column_letter + str(rowD)].style = rowPar

    ws.insert_rows(2)
    ws.insert_rows(3)
    ws.insert_rows(4)

    ws['K3'] = 'OFF TRADE'
    title = ws['K3']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('K3:U3')

    ws['K4'] = 'Promociones'
    title = ws['K4']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('K4:Q4')

    ws['R4'] = 'Iniciativas'
    title = ws['R4']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('R4:U4')

    ws['V3'] = 'ON TRADE'
    title = ws['V3']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('V3:AB3')

    ws['V4'] = 'Promociones'
    title = ws['V4']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('V4:Y4')

    ws['Z4'] = 'Iniciativas'
    title = ws['Z4']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('Z4:AB4')

    ws['AC3'] = 'OXXO'
    title = ws['AC3']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws['AC4'] = 'Iniciativas'
    title = ws['AC4']
    title.font = Font(size=14,bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    title = ws['A6']
    ws.freeze_panes = title

    tab = Table(displayName="Table1",ref="A5:AC"+str(cell_hasta))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    for row in ws.iter_rows('A5:AC5'):
      for cell in row:
        cell.style = headOpe

    nombre_archivo = "84-DECSA MKT Acum Clte Canal Promo Iniciativas-" + datetime.now().date().strftime('%d-%m-%Y') + ".xlsx"
    wb.save(nombre_archivo)

    return send_file('../'+nombre_archivo, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', True, nombre_archivo)

  else:
      return "No hay registros"